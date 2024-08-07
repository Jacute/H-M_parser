import importlib.util
import json

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
import translators as ts
import re
from datetime import datetime

from config.colors import *
from config.config import *
from config.materials import *

from random import random
import shutil
import requests
import os
import logging
import time
import sys
import traceback
import argparse


class Parser:
    def __init__(self):
        self.result = []
        self.count = 0
        
        parser = argparse.ArgumentParser(description='Process some integers.')
        parser.add_argument('--headless', action='store_true', help='headless')
        self.args = parser.parse_args()
        if self.args.headless:
            self.driver = self.get_driver(True)
        else:
            self.driver = self.get_driver(False)

    def get_driver(self, headless):
        try:
            options = webdriver.ChromeOptions()
            if headless:
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')

            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)

            options.add_argument(
                "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")

            # options.add_argument('--disable-dev-shm-usage')
            # options.add_argument('--no-sandbox')
            service = Service(os.path.abspath("chromedriver") if os.name == 'posix' else os.path.abspath("chromedriver.exe"))
            driver = webdriver.Chrome(
                service=service,
                options=options
            )
            driver.set_window_size(1920, 1080)
            driver.implicitly_wait(5)

            self.wait = WebDriverWait(driver, 5)

            return driver
        except Exception as e:
            print('Неудачная настройка браузера!')
            print(traceback.format_exc())
            print(input('Нажмите ENTER, чтобы закрыть эту программу'))
            sys.exit()

    def get_all_products(self):
        self.driver.get(self.CATEGORIE_URL)
        time.sleep(TIMEOUT)
        
        products = []
        while True:
            footer = self.driver.find_element(By.XPATH, '//footer')
            self.driver.execute_script("arguments[0].scrollIntoView();", footer)
            blocks = self.driver.find_elements(By.CLASS_NAME, 'c02f13')
            for block in blocks:
                a = block.find_element(By.XPATH, './/a')
                product_url = a.get_attribute('href')
                if product_url != 'https://www2.hm.com/pl_pl/index.html':
                    products.append(product_url)
            try:
                self.driver.find_element(By.CSS_SELECTOR, '.f05bd4.aa68da.aaa2a2.f8c3c8.ab0e07')
                break
            except:
                pass
            try:
                nextBtn = self.driver.find_element(By.CSS_SELECTOR, '.f05bd4.aaa2a2.ab0e07')
            except:
                break
            self.driver.execute_script("arguments[0].scrollIntoView();", nextBtn)
            time.sleep(1)
            nextBtn.click()
            time.sleep(5)
        products = self.delete_duplicate(products)
        return products

    def parseOne(self, product_url):
        self.driver.get(product_url)
    
        try:
            brand = self.driver.find_element(By.CSS_SELECTOR, '#js-product-name h2').text.strip()
        except Exception:
            brand = 'H&M'
        
        self.driver.execute_script("window.scrollTo(0, 1100)")
        time.sleep(TIMEOUT)

        description_btn = self.driver.find_element(By.XPATH, '//button[@id="toggle-descriptionAccordion"]')
        description_btn.click()
        time.sleep(TIMEOUT)
        description = self.translate(self.driver.find_element(By.CSS_SELECTOR, '.d1cd7b.b475fe.e2b79d').text).strip()

        try:
            material_btn = self.driver.find_element(By.XPATH, '//button[@id="toggle-materialsAndSuppliersAccordion"]')
            material_btn.click()
            time.sleep(TIMEOUT)

            material = self.driver.find_element(By.CSS_SELECTOR, '.d1cd7b.a09145.efef57').text
        except Exception:
            material = 'Bad material'

        try:
            creator_btn = self.driver.find_element(By.CSS_SELECTOR, '.f05bd4.cf896c.c63d19.aaa2a2.d28f9c')
            creator_btn.click()
            time.sleep(TIMEOUT)

            creator = self.translate(self.driver.find_element(By.XPATH, '//div[@class="b4bf73"]/h3').text)
            if creator == 'Инди': creator = 'Индия'
            close = self.driver.find_element(By.XPATH, '//div[@class="f10030"]/button')
            close.click()
            time.sleep(TIMEOUT)
        except:
            creator = 'Bad creator'

        name = self.translate(self.driver.find_element(By.XPATH, '//h1').text).strip()

        price = self.driver.find_element(By.ID, 'product-price').text
        if 'Cena dla Klubowiczów' in price:
            price = price[:price.find('Cena dla Klubowiczów')]
        price = re.findall(r'[0-9 ]+,\d+', price)[0].replace(',', '.').replace(' ', '').strip()
        if price == []: price = re.findall(r'\d+', price)[0].replace(',', '.').replace(' ', '').strip()
        price = self.get_hm_price(price)

        colors = [j.get_attribute('href') for j in self.driver.find_elements(By.XPATH, '//li[@class="list-item"]/a')]
        # if len(colors) >= 8: colors = colors[:7]
        if self.PARSE_TYPE == 'bags':
            if material != 'Bad material' and material != 'N/A 100%' and material.split(' ')[0].lower() != '':
                material = MATERIALS[material.split(' ')[0]]
            else:
                material = ''
            for j in colors:
                try:
                    self.driver.get(j)
                except Exception:
                    continue
                time.sleep(TIMEOUT)

                article_num = re.search('[0-9]{5,}', self.driver.current_url)[0]

                self.wait.until(EC.visibility_of_element_located(
                    (By.XPATH, '//div[@class="product-detail-main-image-container"]/img')))
                main_photo_url = self.driver.find_element(By.XPATH,
                                                            '//div[@class="product-detail-main-image-container"]/img').get_attribute(
                    'src')
                main_photo = self.get_photo(main_photo_url, str(article_num) + '_0.jpeg')

                other_photo_urls = self.driver.find_elements(By.XPATH,
                                                                '//figure[@class="pdp-secondary-image pdp-image"]/img')
                other_photo = ','.join(
                    [self.get_photo(other_photo_urls[i].get_attribute('src'), article_num + '_' + str(i + 1) + '.webp')
                        for i in range(len(other_photo_urls))])

                self.count += 1

                color = self.driver.find_element(By.CLASS_NAME, 'product-input-label').text

                article = 'H&M_' + article_num

                rich = self.RICH.format(name, description, article_num)

                self.COLUMNS['№'] = self.count
                self.COLUMNS['Артикул*'] = article
                self.COLUMNS['Название товара'] = name
                try:
                    self.COLUMNS['Цена, руб.*'] = price
                except:
                    self.COLUMNS['Цена, руб.*'] = 'Bad price'
                self.COLUMNS['Ссылка на главное фото*'] = main_photo
                self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                self.COLUMNS['Название модели (для объединения в одну карточку)*'] = article_num[:-3]
                self.COLUMNS['Цвет товара'] = COLORS[color] if color in COLORS else 'разноцветный'
                self.COLUMNS['Название цвета'] = self.translate(color)
                self.COLUMNS['Страна-изготовитель'] = creator
                self.COLUMNS['Материал'] = material
                self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                self.COLUMNS['Бренд в одежде и обуви*'] = brand
                self.COLUMNS['Rich-контент JSON'] = rich

                self.result.append(self.COLUMNS.copy())
        elif self.PARSE_TYPE == 'clothes':
            main_material = re.split(r' \d{1,3}%', material)[0]
            for j in colors:
                if self.driver.current_url == 'https://www2.hm.com/pl_pl/index.html':
                    continue
                try:
                    self.driver.get(j)
                except Exception:
                    continue
                time.sleep(TIMEOUT)

                article_num = re.search('[0-9]{5,}', self.driver.current_url)[0]

                self.wait.until(EC.visibility_of_element_located(
                    (By.XPATH, '//div[@class="product-detail-main-image-container"]/img')))
                main_photo_url = self.driver.find_element(By.XPATH,
                                                            '//div[@class="product-detail-main-image-container"]/img').get_attribute(
                    'src')
                main_photo = self.get_photo(main_photo_url, str(article_num) + '_0.jpeg')

                other_photo_urls = self.driver.find_elements(By.XPATH,
                                                                '//figure[@class="pdp-secondary-image pdp-image"]/img')
                other_photo = ','.join([self.get_photo(other_photo_urls[i].get_attribute('src'),
                                                        article_num + '_' + str(i + 1) + '.webp') for i in
                                        range(len(other_photo_urls))])

                sizes = self.driver.find_elements(By.XPATH, '//hm-size-selector/ul/li')
                for i in sizes:
                    self.count += 1

                    color = self.driver.find_element(By.CLASS_NAME, 'product-input-label').text

                    size = i.text.split('\n')[0].strip()

                    article = 'H&M_' + article_num + '_' + size

                    rich = self.RICH.format(name, description, article_num)

                    self.COLUMNS['№'] = self.count
                    self.COLUMNS['Артикул*'] = article
                    self.COLUMNS['Название товара'] = name
                    try:
                        self.COLUMNS['Цена, руб.*'] = price
                    except:
                        self.COLUMNS['Цена, руб.*'] = 'Bad price'
                    self.COLUMNS['Ссылка на главное фото*'] = main_photo
                    self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                    self.COLUMNS['Объединить на одной карточке*'] = article_num[:-3]
                    self.COLUMNS['Цвет товара*'] = COLORS[color] if color in COLORS else 'разноцветный'
                    if size.isdigit():
                        self.COLUMNS['Российский размер*'] = str(int(size) + 6)
                    else:
                        try:
                            self.COLUMNS['Российский размер*'] = self.SIZES[size.upper()]
                        except:
                            self.COLUMNS['Российский размер*'] = 'Bad size'  # Если размера нету в таблице размеров
                    self.COLUMNS['Размер производителя'] = size
                    self.COLUMNS['Название цвета'] = self.translate(color)
                    self.COLUMNS['Страна-изготовитель'] = creator
                    self.COLUMNS['Состав материала'] = self.translate(material)
                    self.COLUMNS['Материал'] = self.translate(main_material)
                    self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                    self.COLUMNS['Бренд в одежде и обуви*'] = brand
                    self.COLUMNS['Rich-контент JSON'] = rich

                    self.result.append(self.COLUMNS.copy())
        elif self.PARSE_TYPE == 'shoes':
            material = self.driver.find_element(By.XPATH, '//ul[@class="f94b22"]').text
            main_material, internal_material, sole_material = '', '', ''
            for i in material.split('\n'):
                key, value = i.split(':')
                if key == 'Strona wierzchnia':
                    main_material = MATERIALS[value.split(' ')[0]]
                elif key == 'Podszewka':
                    internal_material = MATERIALS[value.split(' ')[0]]
                elif key == 'Podeszwa zewnętrzna':
                    sole_material = MATERIALS[value.split(' ')[0]]
            for j in colors:
                try:
                    self.driver.get(j)
                except Exception:
                    continue
                time.sleep(TIMEOUT)

                article_num = re.search('[0-9]{5,}', self.driver.current_url)[0]

                self.wait.until(EC.visibility_of_element_located(
                    (By.XPATH, '//div[@class="product-detail-main-image-container"]/img')))
                main_photo_url = self.driver.find_element(By.XPATH,
                                                            '//div[@class="product-detail-main-image-container"]/img').get_attribute(
                    'src')
                main_photo = self.get_photo(main_photo_url, str(article_num) + '_0.jpeg')

                other_photo_urls = self.driver.find_elements(By.XPATH,
                                                                '//figure[@class="pdp-secondary-image pdp-image"]/img')
                other_photo = ','.join([self.get_photo(
                    other_photo_urls[i].get_attribute('src'),
                    article_num + '_' + str(i + 1) + '.webp') for i in
                    range(len(other_photo_urls))])

                sizes = self.driver.find_elements(By.XPATH, '//hm-size-selector/ul/li')
                for i in sizes:
                    self.count += 1

                    color = self.driver.find_element(By.CLASS_NAME, 'product-input-label').text

                    size = i.text.split('\n')[0]

                    article = 'H&M_' + article_num + '_' + size

                    rich = self.RICH.format(name, description, article_num)

                    self.COLUMNS['№'] = self.count
                    self.COLUMNS['Артикул*'] = article
                    self.COLUMNS['Название товара'] = name
                    try:
                        self.COLUMNS['Цена, руб.*'] = price
                    except:
                        self.COLUMNS['Цена, руб.*'] = 'Bad price'
                    self.COLUMNS['Ссылка на главное фото*'] = main_photo
                    self.COLUMNS['Ссылки на дополнительные фото'] = other_photo
                    self.COLUMNS['Объединить на одной карточке*'] = article_num[:-3]
                    self.COLUMNS['Цвет товара*'] = COLORS[color] if color in COLORS else 'разноцветный'
                    try:
                        self.COLUMNS['Российский размер (обуви)*'] = self.SIZES[size.upper()]
                    except:
                        self.COLUMNS['Российский размер (обуви)*'] = 'Bad size'  # Если размера нету в таблице размеров
                    self.COLUMNS['Размер производителя'] = size
                    self.COLUMNS['Название цвета'] = self.translate(color)
                    self.COLUMNS['Страна-изготовитель'] = creator
                    self.COLUMNS['Материал'] = main_material
                    self.COLUMNS['Внутренний материал'] = internal_material
                    self.COLUMNS['Материал подошвы'] = sole_material
                    self.COLUMNS['Таблица размеров JSON'] = self.TABLE_OF_SIZES
                    self.COLUMNS['Бренд в одежде и обуви*'] = brand
                    self.COLUMNS['Rich-контент JSON'] = rich

                    self.result.append(self.COLUMNS.copy())

    def parse(self):
        products = self.get_all_products()
        # products = ['https://www2.hm.com/pl_pl/productpage.1218895002.html']
        for product_url in products[:PARSE_LIMIT]:
            print(f'{products.index(product_url) + 1} of {len(products[:PARSE_LIMIT])}')
            try:
                self.parseOne(product_url)
            except TimeoutException:
                pass
            except Exception:
                self.driver.refresh()
                error = self.driver.current_url + '\n' + traceback.format_exc() + '\n'
                with open('log.log', 'a') as f:
                    f.write(error)

    def check_exists_by_xpath(self, xpath):
        try:
            self.driver.find_element(By.XPATH, xpath)
        except NoSuchElementException:
            return False
        return True

    def get_photo(self, url, name):
        r = requests.get(url, stream=True)
        if r.status_code == 200:
            with open(SAVE_PHOTO_PATH + name, 'wb') as f:
                r.raw.decode_content = True
                shutil.copyfileobj(r.raw, f)
            return 'http://' + HOST + '/H-M_parser/' + SAVE_PHOTO_PATH + name
        else:
            return 'Bad photo'

    def delete_duplicate(self, urls):
        result = []
        set_of_articles = set()
        for url in urls:
            article = re.search(r'[0-9]{5,}', url)[0][:-3]
            if article not in set_of_articles:
                set_of_articles.add(article)
                result.append(url)
        return result

    def translate(self, text):
        while True:
            try:
                result = ts.translate_text(text, to_language='ru', translator='yandex')
                return result
            except:
                pass

    def gPriceDict(self, key):
        return float(PRICE_TABLE[key])

    def get_hm_price(self, pln_price):
        cost_price = ((float(pln_price) / self.gPriceDict("КУРС_USD_ЗЛОТЫ")) * self.gPriceDict("КОЭФ_КОНВЕРТАЦИИ") * self.gPriceDict(
            'КУРС_USD_RUB')) + (self.DELIVERY_PRICE * self.gPriceDict('КУРС_БЕЛ.РУБ_РУБ') * self.gPriceDict(
            'КУРС_EUR_БЕЛ.РУБ'))
        final_price = (cost_price + self.gPriceDict('СРЕД_ЦЕН_ДОСТАВКИ')) / (
                    1 - self.gPriceDict('НАЦЕНКА') - self.OZON_PRICE_MARKUP - self.gPriceDict('ПРОЦЕНТЫ_НАЛОГ') - self.gPriceDict('ПРОЦЕНТЫ_ЭКВАЙРИНГ'))

        final_price = (final_price // 100 + 1) * 100 - 1
        return final_price

    def load_settings(self):
        with open('settings.json', 'r', encoding='utf-8') as f:
            self.settings = json.load(f)
        self.CATEGORIE_URL = self.settings[CATEGORIE]['url']
        self.PARSE_TYPE = self.settings[CATEGORIE]['type_pars']
        self.DELIVERY_PRICE = float(self.settings[CATEGORIE]["ЦЕНА_ДОСТАВКИ_В_КАТЕГОРИИ"])
        self.OZON_PRICE_MARKUP = float(self.settings[CATEGORIE]["ПРОЦЕНТЫ_ОЗОН"])
        self.COLUMNS = self.load_module('columns').COLUMNS
        self.RICH = self.load_module('rich').RICH
        self.SIZES = self.load_module('sizes').SIZES
        self.TABLE_OF_SIZES = self.load_module('table_of_sizes').TABLE_OF_SIZES
        self.MATERIALS = MATERIALS
        self.COLORS = COLORS

    def load_module(self, name):
        spec = importlib.util.spec_from_file_location(name, self.settings[CATEGORIE]['folder_path'] + '/' + name + '.py')
        foo = importlib.util.module_from_spec(spec)
        sys.modules[name] = foo
        spec.loader.exec_module(foo)
        return foo

    def save(self, result):
        wb = load_workbook(filename=f'{self.settings[CATEGORIE]["folder_path"]}/example.xlsx')
        ws = wb['Шаблон']
        alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
        cols = []
        for col in alphabet:
            value = ws[col + '2'].value
            if value:
                cols.append(value)
        for col1 in alphabet:
            for col2 in alphabet:
                value = ws[col1 + col2 + '2'].value
                if value:
                    cols.append(value)

        for row in range(len(result)):
            for col in range(len(cols)):
                if cols[col] not in result[row]:
                    ws.cell(row=4 + row, column=1 + col).value = ''
                else:
                    ws.cell(row=4 + row, column=1 + col).value = result[row][cols[col]]

        wb.save(SAVE_XLSX_PATH + f"{CATEGORIE}_{datetime.now()}.xlsx".replace(':', '.'))

    def start(self):
        try:
            self.load_settings()
            print('--- START PARSING ---')
            self.parse()
            self.save(self.result)
            print('--- END PARSING ---')
        except Exception as e:
            self.save(self.result)
            error = self.driver.current_url + '\n' + traceback.format_exc() + '\n'
            print(error)
            with open('log.log', 'a') as f:
                f.write(error)
        finally:
            self.driver.close()
            self.driver.quit()


def main():
    parser = Parser()
    parser.start()


if __name__ == '__main__':
    if 'photo' not in os.listdir():
        os.mkdir('photo')
    if 'xlsx' not in os.listdir():
        os.mkdir('xlsx')
    if 'log.log' not in os.listdir():
        file = open('log.log', 'w')
        file.close()
    main()

